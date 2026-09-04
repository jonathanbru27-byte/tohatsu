from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
from bson import ObjectId
from openpyxl import Workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('SECRET_KEY', 'tohatsu_dev_fallback_key')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class Motor(BaseModel):
    id: Optional[str] = None
    modelo: str
    potencia: str
    hp_value: int = 0
    tipo: str = ""
    cilindrada: str = ""
    peso_seco: str = ""
    sistema: str = ""
    badge_text: str = "JAPAN TECH"
    caracteristicas: str = ""
    precio: float
    imagen: str
    financiamiento_entrada: float
    financiamiento_cuotas: int = 30

class MotorCreate(BaseModel):
    modelo: str
    potencia: str
    hp_value: int = 0
    tipo: str = ""
    cilindrada: str = ""
    peso_seco: str = ""
    sistema: str = ""
    badge_text: str = "JAPAN TECH"
    caracteristicas: str = ""
    precio: float
    imagen: str
    financiamiento_entrada: float
    financiamiento_cuotas: int = 30

class CalendarioEvento(BaseModel):
    id: Optional[str] = None
    titulo: str = ""
    fecha: str
    hora: str = ""
    localidad: str
    descripcion: str

class CalendarioEventoCreate(BaseModel):
    titulo: str = ""
    fecha: str
    hora: str = ""
    localidad: str
    descripcion: str

class Configuracion(BaseModel):
    whatsapp_ventas: str
    whatsapp_repuestos: str
    whatsapp_servicio: str

class Repuesto(BaseModel):
    id: Optional[str] = None
    nombre: str
    descripcion: str
    precio: float
    imagen: str = ""
    categoria: str = "General"
    stock: int = 0
    modelos_compatibles: str = ""
    codigo: str = ""

class RepuestoCreate(BaseModel):
    nombre: str
    descripcion: str
    precio: float
    imagen: str = ""
    categoria: str = "General"
    stock: int = 0
    modelos_compatibles: str = ""
    codigo: str = ""

class Asesor(BaseModel):
    id: Optional[str] = None
    nombre: str
    whatsapp: str
    provincia: str  # Manabí, Santa Elena, Esmeraldas, Guayas

class AsesorCreate(BaseModel):
    nombre: str
    whatsapp: str
    provincia: str

class Lead(BaseModel):
    id: Optional[str] = None
    fecha: str
    hora: str
    nombre: str = ""
    telefono: str = ""
    provincia: str = ""
    interes: str  # motor / repuesto / servicio
    detalle: str = ""

class LeadCreate(BaseModel):
    nombre: str = ""
    telefono: str = ""
    provincia: str = ""
    interes: str
    detalle: str = ""

class AdminLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

# ==================== AUTH ====================

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudo validar las credenciales",
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = await db.admins.find_one({"username": username})
    if user is None:
        raise credentials_exception
    return user

# ==================== SEED ====================

async def seed_initial_data():
    admin_exists = await db.admins.find_one({"username": "admin"})
    if not admin_exists:
        await db.admins.insert_one({
            "username": "admin",
            "password": get_password_hash("admin123")
        })
        logger.info("Admin user created")

    config_exists = await db.configuracion.find_one({})
    if not config_exists:
        await db.configuracion.insert_one({
            "whatsapp_ventas": "593999999999",
            "whatsapp_repuestos": "593988888888",
            "whatsapp_servicio": "593977777777"
        })

    # Seed repuestos
    repuestos_count = await db.repuestos.count_documents({})
    if repuestos_count == 0:
        sample_repuestos = [
            {"nombre": "Filtro de aceite OEM", "descripcion": "Filtro original Tohatsu para motores 4 tiempos 40-90 HP. Garantiza máxima protección del motor.", "precio": 28.50, "imagen": "", "categoria": "Filtros", "stock": 25},
            {"nombre": "Bujía NGK BR8HS", "descripcion": "Bujía original para motores Tohatsu de 2 y 4 tiempos. Vendida por unidad.", "precio": 12.00, "imagen": "", "categoria": "Encendido", "stock": 50},
            {"nombre": "Impulsor de bomba de agua", "descripcion": "Impulsor de goma compatible con motores 9.8-30 HP. Reemplazo recomendado cada temporada.", "precio": 45.00, "imagen": "", "categoria": "Refrigeración", "stock": 30},
            {"nombre": "Aceite 4T Tohatsu 10W-40", "descripcion": "Aceite sintético específico para motores fuera de borda 4 tiempos. Envase 1L.", "precio": 22.00, "imagen": "", "categoria": "Lubricantes", "stock": 40},
            {"nombre": "Hélice Aluminio 3 palas 10x12", "descripcion": "Hélice estándar para motores 25-50 HP. Diámetro 10\", paso 12\". Material aluminio reforzado.", "precio": 185.00, "imagen": "", "categoria": "Hélices", "stock": 15},
            {"nombre": "Ánodo de zinc sacrificial", "descripcion": "Ánodo de protección anticorrosiva para motores 25-115 HP. Protege componentes metálicos.", "precio": 18.50, "imagen": "", "categoria": "Protección", "stock": 35},
            {"nombre": "Kit reparación carburador", "descripcion": "Kit completo con juntas, agujas y diafragmas para carburador motores 9.8-30 HP.", "precio": 65.00, "imagen": "", "categoria": "Combustible", "stock": 12},
            {"nombre": "Cable acelerador 5m", "descripcion": "Cable de acelerador universal de 5 metros con terminales. Compatible con controles remotos.", "precio": 42.00, "imagen": "", "categoria": "Mandos", "stock": 20},
            {"nombre": "Termostato motor", "descripcion": "Termostato original 50°C para motores Tohatsu 4 tiempos 40-90 HP. Regula temperatura óptima.", "precio": 38.00, "imagen": "", "categoria": "Refrigeración", "stock": 18},
            {"nombre": "Manguera combustible + bulbo", "descripcion": "Manguera de combustible 2.5m con bulbo cebador y conector rápido. Universal Tohatsu.", "precio": 32.00, "imagen": "", "categoria": "Combustible", "stock": 28},
        ]
        await db.repuestos.insert_many(sample_repuestos)
        logger.info("Repuestos seed created")

# ==================== ENDPOINTS ====================

@api_router.get("/")
async def root():
    return {"message": "Tohatsu Motors API"}

@api_router.post("/auth/login", response_model=Token)
async def login(admin: AdminLogin):
    user = await db.admins.find_one({"username": admin.username})
    if not user or not verify_password(admin.password, user["password"]):
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    access_token = create_access_token(data={"sub": user["username"]})
    return {"access_token": access_token, "token_type": "bearer"}

def motor_to_dict(motor):
    return Motor(id=str(motor["_id"]), **{k: v for k, v in motor.items() if k != "_id"})

@api_router.get("/motors", response_model=List[Motor])
async def get_motors():
    motors = await db.motors.find().sort("hp_value", -1).to_list(1000)
    return [motor_to_dict(m) for m in motors]

@api_router.get("/motors/{motor_id}", response_model=Motor)
async def get_motor(motor_id: str):
    try:
        motor = await db.motors.find_one({"_id": ObjectId(motor_id)})
        if not motor:
            raise HTTPException(status_code=404, detail="Motor no encontrado")
        return motor_to_dict(motor)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/motors", response_model=Motor)
async def create_motor(motor: MotorCreate, current_user: dict = Depends(get_current_user)):
    motor_dict = motor.dict()
    result = await db.motors.insert_one(motor_dict)
    motor_dict["id"] = str(result.inserted_id)
    return Motor(**motor_dict)

@api_router.put("/motors/{motor_id}", response_model=Motor)
async def update_motor(motor_id: str, motor: MotorCreate, current_user: dict = Depends(get_current_user)):
    try:
        motor_dict = motor.dict()
        await db.motors.update_one({"_id": ObjectId(motor_id)}, {"$set": motor_dict})
        motor_dict["id"] = motor_id
        return Motor(**motor_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/motors/{motor_id}")
async def delete_motor(motor_id: str, current_user: dict = Depends(get_current_user)):
    try:
        result = await db.motors.delete_one({"_id": ObjectId(motor_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Motor no encontrado")
        return {"message": "Motor eliminado"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/calendar", response_model=List[CalendarioEvento])
async def get_calendar():
    eventos = await db.calendario.find().sort("fecha", 1).to_list(1000)
    return [CalendarioEvento(id=str(e["_id"]), **{k: v for k, v in e.items() if k != "_id"}) for e in eventos]

@api_router.post("/calendar", response_model=CalendarioEvento)
async def create_evento(evento: CalendarioEventoCreate, current_user: dict = Depends(get_current_user)):
    evento_dict = evento.dict()
    result = await db.calendario.insert_one(evento_dict)
    evento_dict["id"] = str(result.inserted_id)
    return CalendarioEvento(**evento_dict)

@api_router.put("/calendar/{evento_id}", response_model=CalendarioEvento)
async def update_evento(evento_id: str, evento: CalendarioEventoCreate, current_user: dict = Depends(get_current_user)):
    try:
        evento_dict = evento.dict()
        await db.calendario.update_one({"_id": ObjectId(evento_id)}, {"$set": evento_dict})
        evento_dict["id"] = evento_id
        return CalendarioEvento(**evento_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/calendar/{evento_id}")
async def delete_evento(evento_id: str, current_user: dict = Depends(get_current_user)):
    try:
        result = await db.calendario.delete_one({"_id": ObjectId(evento_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Evento no encontrado")
        return {"message": "Evento eliminado"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/config", response_model=Configuracion)
async def get_config():
    config = await db.configuracion.find_one({})
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    return Configuracion(**{k: v for k, v in config.items() if k != "_id"})

@api_router.put("/config", response_model=Configuracion)
async def update_config(config: Configuracion, current_user: dict = Depends(get_current_user)):
    config_dict = config.dict()
    await db.configuracion.update_one({}, {"$set": config_dict}, upsert=True)
    return config

# ==================== REPUESTOS ====================

def repuesto_to_dict(r):
    return Repuesto(id=str(r["_id"]), **{k: v for k, v in r.items() if k != "_id"})

@api_router.get("/repuestos", response_model=List[Repuesto])
async def get_repuestos():
    repuestos = await db.repuestos.find().to_list(1000)
    return [repuesto_to_dict(r) for r in repuestos]

@api_router.get("/repuestos/{repuesto_id}", response_model=Repuesto)
async def get_repuesto(repuesto_id: str):
    try:
        r = await db.repuestos.find_one({"_id": ObjectId(repuesto_id)})
        if not r:
            raise HTTPException(status_code=404, detail="Repuesto no encontrado")
        return repuesto_to_dict(r)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/repuestos", response_model=Repuesto)
async def create_repuesto(repuesto: RepuestoCreate, current_user: dict = Depends(get_current_user)):
    try:
        rd = repuesto.dict()
        result = await db.repuestos.insert_one(rd)
        rd["id"] = str(result.inserted_id)
        return Repuesto(**rd)
    except Exception as e:
        print("===== ERROR AL CREAR REPUESTO =====")
        print(f"Mensaje: {str(e)}")
        traceback.print_exc()
        print("===================================")
        raise HTTPException(status_code=500, detail=f"Error guardando repuesto: {str(e)}")

@api_router.put("/repuestos/{repuesto_id}", response_model=Repuesto)
async def update_repuesto(repuesto_id: str, repuesto: RepuestoCreate, current_user: dict = Depends(get_current_user)):
    try:
        rd = repuesto.dict()
        await db.repuestos.update_one({"_id": ObjectId(repuesto_id)}, {"$set": rd})
        rd["id"] = repuesto_id
        return Repuesto(**rd)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/repuestos/{repuesto_id}")
async def delete_repuesto(repuesto_id: str, current_user: dict = Depends(get_current_user)):
    try:
        result = await db.repuestos.delete_one({"_id": ObjectId(repuesto_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Repuesto no encontrado")
        return {"message": "Repuesto eliminado"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== ASESORES ====================

def asesor_to_dict(a):
    return Asesor(id=str(a["_id"]), **{k: v for k, v in a.items() if k != "_id"})

@api_router.get("/asesores", response_model=List[Asesor])
async def get_asesores():
    asesores = await db.asesores.find().to_list(1000)
    return [asesor_to_dict(a) for a in asesores]

@api_router.post("/asesores", response_model=Asesor)
async def create_asesor(asesor: AsesorCreate, current_user: dict = Depends(get_current_user)):
    ad = asesor.dict()
    result = await db.asesores.insert_one(ad)
    ad["id"] = str(result.inserted_id)
    return Asesor(**ad)

@api_router.put("/asesores/{asesor_id}", response_model=Asesor)
async def update_asesor(asesor_id: str, asesor: AsesorCreate, current_user: dict = Depends(get_current_user)):
    try:
        ad = asesor.dict()
        await db.asesores.update_one({"_id": ObjectId(asesor_id)}, {"$set": ad})
        ad["id"] = asesor_id
        return Asesor(**ad)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/asesores/{asesor_id}")
async def delete_asesor(asesor_id: str, current_user: dict = Depends(get_current_user)):
    try:
        result = await db.asesores.delete_one({"_id": ObjectId(asesor_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Asesor no encontrado")
        return {"message": "Asesor eliminado"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/asesores/by-provincia/{provincia}")
async def get_asesor_by_provincia(provincia: str):
    asesor = await db.asesores.find_one({"provincia": provincia})
    if not asesor:
        # Fallback al número general de ventas
        config = await db.configuracion.find_one({})
        if config:
            return {"nombre": "Asesor General", "whatsapp": config.get("whatsapp_ventas", ""), "provincia": "General", "is_general": True}
        raise HTTPException(status_code=404, detail="No hay asesor disponible")
    return {**asesor_to_dict(asesor).dict(), "is_general": False}

# ==================== LEADS ====================

@api_router.get("/leads", response_model=List[Lead])
async def get_leads(current_user: dict = Depends(get_current_user)):
    leads = await db.leads.find().sort("fecha", -1).to_list(10000)
    return [Lead(id=str(l["_id"]), **{k: v for k, v in l.items() if k != "_id"}) for l in leads]

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead: LeadCreate):
    now = datetime.now()
    lead_dict = lead.dict()
    lead_dict["fecha"] = now.strftime("%Y-%m-%d")
    lead_dict["hora"] = now.strftime("%H:%M:%S")
    result = await db.leads.insert_one(lead_dict)
    lead_dict["id"] = str(result.inserted_id)
    return Lead(**lead_dict)

@api_router.get("/leads/export/xlsx")
async def export_leads_xlsx(current_user: dict = Depends(get_current_user)):
    leads = await db.leads.find().sort("fecha", -1).to_list(100000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Tohatsu"
    headers = ["Fecha", "Hora", "Nombre", "Teléfono", "Provincia", "Interés", "Detalle"]
    ws.append(headers)
    # Estilo header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for l in leads:
        ws.append([
            l.get("fecha", ""),
            l.get("hora", ""),
            l.get("nombre", ""),
            l.get("telefono", ""),
            l.get("provincia", ""),
            l.get("interes", ""),
            l.get("detalle", ""),
        ])
    # Ancho de columnas
    widths = [12, 10, 22, 18, 16, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads_tohatsu.xlsx"},
    )

# ==================== SEED ENDPOINT ====================

# Catálogo completo de motores Tohatsu
TOHATSU_MOTORS = [
    {
        "modelo": "Tohatsu M3.5B",
        "potencia": "3.5 HP",
        "hp_value": 3,
        "tipo": "2 Tiempos",
        "cilindrada": "74.6 cc",
        "peso_seco": "13 kg",
        "sistema": "Arranque manual, CDI",
        "badge_text": "PORTÁTIL",
        "caracteristicas": "Motor ultraligero ideal para botes inflables y pequeñas embarcaciones. Sistema de enfriamiento por agua. Tanque de combustible integrado de 1.0L. Altura de popa corta (381mm). Ideal para pesca en lagos y ríos.",
        "precio": 1250.00,
        "imagen": "https://images.pexels.com/photos/28170856/pexels-photo-28170856.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 250.00,
        "financiamiento_cuotas": 24
    },
    {
        "modelo": "Tohatsu MFS6",
        "potencia": "6 HP",
        "hp_value": 6,
        "tipo": "4 Tiempos",
        "cilindrada": "138 cc",
        "peso_seco": "26 kg",
        "sistema": "Arranque manual, EFI disponible",
        "badge_text": "ECO FRIENDLY",
        "caracteristicas": "Motor de 4 tiempos silencioso y económico. Bajo consumo de combustible. Sistema de lubricación por cárter húmedo. Tanque externo de 12L incluido. Altura de popa corta/larga disponible. Perfecto para veleros y botes auxiliares.",
        "precio": 2150.00,
        "imagen": "https://images.pexels.com/photos/30094170/pexels-photo-30094170.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 430.00,
        "financiamiento_cuotas": 30
    },
    {
        "modelo": "Tohatsu MFS9.9",
        "potencia": "9.9 HP",
        "hp_value": 9,
        "tipo": "4 Tiempos",
        "cilindrada": "209 cc",
        "peso_seco": "38 kg",
        "sistema": "Arranque manual/eléctrico, EFI",
        "badge_text": "BEST SELLER",
        "caracteristicas": "El motor más vendido de la gama. Inyección electrónica EFI para máxima eficiencia. Alternador de 12V/6A. Sistema de alerta de sobrecalentamiento. Ideal para botes de aluminio y fibra hasta 4m. No requiere licencia náutica en muchos países.",
        "precio": 3200.00,
        "imagen": "https://images.pexels.com/photos/9592461/pexels-photo-9592461.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 640.00,
        "financiamiento_cuotas": 30
    },
    {
        "modelo": "Tohatsu M9.8",
        "potencia": "9.8 HP",
        "hp_value": 9,
        "tipo": "2 Tiempos",
        "cilindrada": "169 cc",
        "peso_seco": "26 kg",
        "sistema": "Arranque manual, CDI",
        "badge_text": "CLÁSICO",
        "caracteristicas": "Motor 2 tiempos de probada confiabilidad. Relación peso-potencia excepcional. Sistema de mezcla de aceite automático opcional. Diseño compacto y fácil mantenimiento. Popular entre pescadores profesionales.",
        "precio": 2650.00,
        "imagen": "https://images.pexels.com/photos/16135856/pexels-photo-16135856.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 530.00,
        "financiamiento_cuotas": 30
    },
    {
        "modelo": "Tohatsu MX18",
        "potencia": "18 HP",
        "hp_value": 18,
        "tipo": "4 Tiempos EFI",
        "cilindrada": "333 cc",
        "peso_seco": "45 kg",
        "sistema": "Arranque eléctrico, EFI, Power Trim",
        "badge_text": "NUEVO 2025",
        "caracteristicas": "Nueva generación de motores Tohatsu. Sistema EFI de última generación. Power Trim & Tilt eléctrico. Alternador de 12V/12A. Sistema de diagnóstico digital. Garantía extendida de 5 años. Ideal para botes de pesca deportiva.",
        "precio": 4350.00,
        "imagen": "https://images.pexels.com/photos/32967413/pexels-photo-32967413.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 870.00,
        "financiamiento_cuotas": 36
    },
    {
        "modelo": "Tohatsu MX30",
        "potencia": "30 HP",
        "hp_value": 30,
        "tipo": "4 Tiempos EFI",
        "cilindrada": "526 cc",
        "peso_seco": "72 kg",
        "sistema": "Arranque eléctrico, EFI, Power Trim",
        "badge_text": "JAPAN TECH",
        "caracteristicas": "Motor de media potencia con tecnología japonesa de punta. Inyección multipunto. Sistema de enfriamiento de doble circuito. Caja de cambios Forward-Neutral-Reverse. Compatible con control remoto. Excelente para lanchas de hasta 5.5m.",
        "precio": 5850.00,
        "imagen": "https://images.pexels.com/photos/14815453/pexels-photo-14815453.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 1170.00,
        "financiamiento_cuotas": 36
    },
    {
        "modelo": "Tohatsu MFS30",
        "potencia": "30 HP",
        "hp_value": 30,
        "tipo": "4 Tiempos",
        "cilindrada": "526 cc",
        "peso_seco": "68 kg",
        "sistema": "Arranque manual/eléctrico, Carburador",
        "badge_text": "CONFIABLE",
        "caracteristicas": "Motor 4 tiempos de la serie MFS. Sistema de carburador de alto rendimiento. Excelente relación calidad-precio. Bajo mantenimiento. Alternador de 12V/10A. Power Tilt manual. Ideal para uso recreativo y pesca artesanal.",
        "precio": 5200.00,
        "imagen": "https://images.pexels.com/photos/847393/pexels-photo-847393.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 1040.00,
        "financiamiento_cuotas": 36
    },
    {
        "modelo": "Tohatsu MX40",
        "potencia": "40 HP",
        "hp_value": 40,
        "tipo": "4 Tiempos EFI",
        "cilindrada": "866 cc",
        "peso_seco": "87 kg",
        "sistema": "Arranque eléctrico, EFI, Power Trim & Tilt",
        "badge_text": "PRO SERIES",
        "caracteristicas": "Motor profesional para uso intensivo. 3 cilindros en línea. Sistema TLDI de inyección directa disponible. Alternador de alta capacidad 12V/18A. Sistema de alerta múltiple (temperatura, presión de aceite, RPM). Ideal para pesca comercial y turismo náutico.",
        "precio": 7950.00,
        "imagen": "https://images.pexels.com/photos/8669061/pexels-photo-8669061.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 1590.00,
        "financiamiento_cuotas": 48
    },
    {
        "modelo": "Tohatsu MX50",
        "potencia": "50 HP",
        "hp_value": 50,
        "tipo": "4 Tiempos EFI",
        "cilindrada": "866 cc",
        "peso_seco": "93 kg",
        "sistema": "Arranque eléctrico, EFI, Power Trim & Tilt, TLDI",
        "badge_text": "PREMIUM",
        "caracteristicas": "Tope de gama de la serie MX. Máxima potencia y eficiencia. Sistema TLDI de inyección directa estratificada. Menor consumo y emisiones. Tecnología Variable Valve Timing. Control digital con pantalla multifunción opcional. Garantía premium de 5 años. Para embarcaciones de hasta 6.5m.",
        "precio": 9500.00,
        "imagen": "https://images.pexels.com/photos/9381649/pexels-photo-9381649.jpeg?auto=compress&cs=tinysrgb&w=600",
        "financiamiento_entrada": 1900.00,
        "financiamiento_cuotas": 48
    }
]

SEED_KEY = os.environ.get('SEED_KEY', 'tohatsu2025seed')

@api_router.post("/seed")
async def run_seed(key: str = None):
    """Ejecuta el seed de la base de datos. Requiere clave de seguridad."""
    if key != SEED_KEY:
        raise HTTPException(status_code=403, detail="Clave de seed inválida")
    
    try:
        # Seed motors
        await db.motors.delete_many({})
        result = await db.motors.insert_many(TOHATSU_MOTORS)
        motors_count = len(result.inserted_ids)
        
        # Seed admin
        admin = await db.admins.find_one({"username": "admin"})
        admin_created = False
        if not admin:
            await db.admins.insert_one({
                "username": "admin",
                "password": pwd_context.hash("admin123")
            })
            admin_created = True
        
        # Seed config
        config = await db.configuracion.find_one({})
        config_created = False
        if not config:
            await db.configuracion.insert_one({
                "whatsapp_ventas": "593999999999",
                "whatsapp_repuestos": "593988888888",
                "whatsapp_servicio": "593977777777"
            })
            config_created = True
        
        return {
            "success": True,
            "message": "Seed ejecutado correctamente",
            "motors_inserted": motors_count,
            "admin_created": admin_created,
            "config_created": config_created
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error ejecutando seed: {str(e)}")

@api_router.get("/seed")
async def seed_info():
    """Información sobre el endpoint de seed"""
    return {
        "endpoint": "/api/seed",
        "method": "POST",
        "description": "Ejecuta el seed de la base de datos",
        "usage": "POST /api/seed?key=TU_CLAVE_DE_SEED",
        "note": "La clave por defecto es 'tohatsu2025seed' o configura SEED_KEY en las variables de entorno"
    }
import traceback
import requests

class ImageUploadRequest(BaseModel):
    image: str

@api_router.post("/upload-image")
async def upload_image(data: ImageUploadRequest):
    try:
        """
        Recibe una imagen en base64, la sube a ImgBB y devuelve la URL pública.
        """
        IMGBB_API_KEY = os.environ.get("IMGBB_API_KEY", "")
        
        # Agregamos un print para que Render lo muestre en la consola
        if not IMGBB_API_KEY:
            print("===== ERROR: FALTA IMGBB_API_KEY EN RENDER =====")
            raise HTTPException(status_code=500, detail="IMGBB_API_KEY no configurada en el servidor")
        
        base64_string = data.image
        if "," in base64_string:
            b64_clean = base64_string.split(",", 1)[1]
        else:
            b64_clean = base64_string
        
        url = "https://api.imgbb.com/1/upload"
        payload = {"key": IMGBB_API_KEY, "image": b64_clean}
        
        response = requests.post(url, data=payload, timeout=60)
        response.raise_for_status()
        result = response.json()
        
        if result.get("success"):
            return {"url": result["data"]["url"]}
        else:
            error_msg = result.get("error", {}).get("message", "Error desconocido de ImgBB")
            print(f"===== ERROR DE IMGBB: {error_msg} =====")
            raise HTTPException(status_code=400, detail=error_msg)
            
    except Exception as e:
        # AQUÍ CAPTURAMOS Y MOSTRAMOS CUALQUIER OTRO ERROR REAL EN LA CONSOLA DE RENDER
        print("===== ERROR CRÍTICO EN UPLOAD-IMAGE =====")
        print(f"Mensaje: {str(e)}")
        traceback.print_exc()
        print("=========================================")
        raise HTTPException(status_code=500, detail=f"Error subiendo imagen: {str(e)}")
app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await seed_initial_data()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
