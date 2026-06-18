"""Tests for /api/leads CRUD + xlsx export."""
import re
import zipfile
import io
import pytest
from datetime import datetime


class TestLeadsAuth:
    def test_get_leads_requires_auth(self, api_client, base_url):
        r = api_client.get(f"{base_url}/api/leads", timeout=15)
        assert r.status_code in (401, 403)

    def test_export_xlsx_requires_auth(self, api_client, base_url):
        r = api_client.get(f"{base_url}/api/leads/export/xlsx", timeout=15)
        assert r.status_code in (401, 403)


class TestLeadsCreate:
    def test_create_lead_no_auth_required(self, api_client, base_url):
        payload = {
            "nombre": "TEST_Cliente_Lead",
            "telefono": "0991234567",
            "provincia": "Manabí",
            "interes": "motor",
            "detalle": "Interesado en Tohatsu 40 HP",
        }
        r = api_client.post(f"{base_url}/api/leads", json=payload, timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["nombre"] == payload["nombre"]
        assert data["telefono"] == payload["telefono"]
        assert data["provincia"] == payload["provincia"]
        assert data["interes"] == payload["interes"]
        assert data["detalle"] == payload["detalle"]
        # Server must auto-fill fecha YYYY-MM-DD and hora HH:MM:SS
        assert re.match(r"^\d{4}-\d{2}-\d{2}$", data["fecha"]), data["fecha"]
        assert re.match(r"^\d{2}:\d{2}:\d{2}$", data["hora"]), data["hora"]
        # Fecha should be today's date (server local time)
        today = datetime.now().strftime("%Y-%m-%d")
        assert data["fecha"] == today
        assert data.get("id")

    def test_create_lead_interes_repuesto(self, api_client, base_url):
        r = api_client.post(
            f"{base_url}/api/leads",
            json={"nombre": "TEST_R", "telefono": "0", "provincia": "Guayas", "interes": "repuesto", "detalle": ""},
            timeout=15,
        )
        assert r.status_code == 200
        assert r.json()["interes"] == "repuesto"

    def test_create_lead_interes_servicio(self, api_client, base_url):
        r = api_client.post(
            f"{base_url}/api/leads",
            json={"nombre": "TEST_S", "telefono": "0", "provincia": "El Oro", "interes": "servicio", "detalle": ""},
            timeout=15,
        )
        assert r.status_code == 200
        assert r.json()["interes"] == "servicio"


class TestLeadsList:
    def test_get_leads_with_auth(self, api_client, base_url, auth_headers):
        r = api_client.get(f"{base_url}/api/leads", headers=auth_headers, timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        assert isinstance(data, list)
        # We created at least 3 leads earlier in this session
        assert len(data) >= 3
        # Verify shape
        sample = data[0]
        for key in ("id", "fecha", "hora", "nombre", "telefono", "provincia", "interes", "detalle"):
            assert key in sample


class TestLeadsExportXlsx:
    def test_export_xlsx_content(self, api_client, base_url, auth_headers):
        r = api_client.get(
            f"{base_url}/api/leads/export/xlsx", headers=auth_headers, timeout=30
        )
        assert r.status_code == 200, r.text[:500]

        # Content-Type
        ctype = r.headers.get("Content-Type", "")
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in ctype, ctype

        # Content-Disposition attachment
        cdisp = r.headers.get("Content-Disposition", "")
        assert "attachment" in cdisp.lower(), cdisp
        assert ".xlsx" in cdisp.lower()

        # PK zip header
        assert r.content[:2] == b"PK", "Body is not a valid xlsx (PK zip header missing)"

        # Validate file structure and headers row using openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 1
        headers_row = list(rows[0])
        expected = ["Fecha", "Hora", "Nombre", "Teléfono", "Provincia", "Interés", "Detalle"]
        assert headers_row == expected, headers_row
        # At least one data row
        assert len(rows) >= 2
